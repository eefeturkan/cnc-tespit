"""
CNC Parça Ölçüm Sistemi — Rapor Oluşturucu
PDF ve Excel formatında ölçüm raporu üretir.
"""

import io
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===================================================================
# Font Tanımlamaları (Türkçe Karakter Desteği için)
# ===================================================================
import os
font_dir = Path(__file__).parent / "assets" / "fonts"
try:
    pdfmetrics.registerFont(TTFont('Roboto', str(font_dir / 'Roboto-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('Roboto-Bold', str(font_dir / 'Roboto-Bold.ttf')))
    DEFAULT_FONT = 'Roboto'
    BOLD_FONT = 'Roboto-Bold'
except Exception as e:
    print(f"Uyarı: Roboto fontu yüklenemedi, Helvetica kullanılacak. Hata: {e}")
    DEFAULT_FONT = 'Helvetica'
    BOLD_FONT = 'Helvetica-Bold'

# ===================================================================
# Renkler
# ===================================================================
DARK_BG = HexColor("#0f172a") # Slate 900
HEADER_BG = HexColor("#1e293b") # Slate 800
ACCENT = HexColor("#2563eb") # Blue 600
ROW_ALT = HexColor("#f8fafc") # Slate 50
TEXT_PRIMARY = HexColor("#0f172a")
TEXT_SECONDARY = HexColor("#64748b")
SUCCESS = HexColor("#16a34a") # Green 600
DANGER = HexColor("#dc2626") # Red 600
BORDER_COLOR = HexColor("#cbd5e1") # Slate 300

# ===================================================================
# PDF Base Header/Footer Canvas
# ===================================================================
def create_report_canvas(canvas, doc):
    canvas.saveState()
    # Footer
    canvas.setFont(DEFAULT_FONT, 8)
    canvas.setFillColor(TEXT_SECONDARY)
    canvas.drawString(15*mm, 10*mm, "NEXORA® CNC Kalite Kontrol Sistemi")
    canvas.drawRightString(200*mm, 10*mm, f"Sayfa {doc.page}")
    
    # Header Line
    canvas.setStrokeColor(BORDER_COLOR)
    canvas.setLineWidth(0.5)
    canvas.line(15*mm, 285*mm, 200*mm, 285*mm)
    canvas.restoreState()

# ===================================================================
# PDF Rapor
# ===================================================================
def generate_pdf_report(
    measurement_table: List[Dict],
    summary: Dict,
    calibration_info: Dict,
    image_path: Optional[str] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """
    PDF ölçüm raporu oluşturur.

    Returns:
        PDF dosyasının byte içeriği
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=25 * mm, bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=20, textColor=DARK_BG, spaceAfter=2 * mm,
        fontName=BOLD_FONT, alignment=TA_LEFT
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=10, textColor=TEXT_SECONDARY,
        spaceAfter=8 * mm, fontName=DEFAULT_FONT, alignment=TA_LEFT
    )
    section_style = ParagraphStyle(
        "SectionTitle", parent=styles["Heading2"],
        fontSize=13, textColor=HEADER_BG, spaceBefore=8 * mm,
        spaceAfter=4 * mm, fontName=BOLD_FONT,
        borderPadding=(0,0,2,0),
        borderColor=ACCENT, borderWidth=1 # Alt Çizgi
    )
    normal_style = ParagraphStyle(
        "NormalText", parent=styles["Normal"],
        fontSize=10, textColor=TEXT_PRIMARY, fontName=DEFAULT_FONT
    )

    elements = []

    # Başlık
    elements.append(Paragraph("CNC Parça Ölçüm Raporu", title_style))
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    elements.append(Paragraph(f"Oluşturulma: {now}", subtitle_style))

    # Kalibrasyon bilgisi
    elements.append(Paragraph("Kalibrasyon Bilgileri", section_style))
    cal_data = [
        ["Parametre", "Değer"],
        ["Piksel/mm", f"{calibration_info.get('pixels_per_mm', 1.0):.4f}"],
        ["Kalibrasyon Adı", calibration_info.get("name", "Varsayılan")],
    ]
    cal_table = Table(cal_data, colWidths=[60 * mm, 80 * mm])
    cal_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(cal_table)
    elements.append(Spacer(1, 4 * mm))

    # Özet
    elements.append(Paragraph("Ölçüm Özeti", section_style))
    if summary.get("total_sections", 0) > 0:
        summary_data = [
            ["Parametre", "Değer"],
            ["Toplam Bölüm", str(summary["total_sections"])],
            ["Min Çap", f"{summary['min_diameter_mm']:.4f} mm"],
            ["Max Çap", f"{summary['max_diameter_mm']:.4f} mm"],
            ["Toplam Uzunluk", f"{summary['total_length_mm']:.4f} mm"],
        ]
    else:
        summary_data = [
            ["Parametre", "Değer"],
            ["Toplam Bölüm", "0"],
        ]
    sum_table = Table(summary_data, colWidths=[60 * mm, 80 * mm])
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(sum_table)
    elements.append(Spacer(1, 6 * mm))

    # Ölçüm tablosu
    elements.append(Paragraph("Ölçüm Sonuçları", section_style))
    table_data = [["ID", "Tip", "Açıklama", "Ölçülen (mm)", "Hedef", "Tolerans", "Durum"]]
    
    # Hücre stili komutları
    pass_fail_styles = []
    
    for idx, row in enumerate(measurement_table):
        target_str = str(row.get('target')) if row.get('target') is not None else "-"
        tol_str = f"±{row.get('tol')}" if row.get('tol') is not None else "-"
        status = row.get('status') or "-"
        
        table_data.append([
            row["id"],
            row["type"],
            row["description"],
            f"{row['measured_mm']:.4f}",
            target_str,
            tol_str,
            status
        ])
        
        # Sütun endeksleri: ID=0, Tip=1, Açıklama=2, Ölçülen=3, Hedef=4, Tolerans=5, Durum=6
        current_row_idx = idx + 1 # Başlık satırından sonra
        
        # Rozet (Badge) Görünümü
        if status == "PASS":
            pass_fail_styles.append(("BACKGROUND", (6, current_row_idx), (6, current_row_idx), SUCCESS))
            pass_fail_styles.append(("TEXTCOLOR", (6, current_row_idx), (6, current_row_idx), white))
            pass_fail_styles.append(("FONTNAME", (6, current_row_idx), (6, current_row_idx), BOLD_FONT))
        elif status == "FAIL":
            pass_fail_styles.append(("BACKGROUND", (6, current_row_idx), (6, current_row_idx), DANGER))
            pass_fail_styles.append(("TEXTCOLOR", (6, current_row_idx), (6, current_row_idx), white))
            pass_fail_styles.append(("FONTNAME", (6, current_row_idx), (6, current_row_idx), BOLD_FONT))

    # Yeni genişlik (A4: 190 mm)
    col_widths = [12 * mm, 15 * mm, 55 * mm, 25 * mm, 20 * mm, 20 * mm, 20 * mm]
    measure_table = Table(table_data, colWidths=col_widths)

    # Tablo stili
    table_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG), # Koyu başlık
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), BOLD_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ALIGN", (3, 1), (6, -1), "CENTER"), # Tüm sayısal kolonları sınırla
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), # Hücre içi dikey hizalama
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("FONTNAME", (0, 1), (-1, -1), DEFAULT_FONT) # Gövde metinleri
    ]
    # Satır renklendirme
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style_cmds.append(("BACKGROUND", (0, i), (-1, i), ROW_ALT))

    table_style_cmds.extend(pass_fail_styles)
    measure_table.setStyle(TableStyle(table_style_cmds))
    elements.append(measure_table)

    # Görüntü varsa ekle
    if image_path and os.path.exists(image_path):
        elements.append(Spacer(1, 8 * mm))
        elements.append(Paragraph("Ölçüm Görüntüsü", section_style))
        try:
            img = RLImage(image_path)
            # A4 genişliğine sığdır
            max_w = 170 * mm
            ratio = img.imageWidth / img.imageHeight
            img_w = min(max_w, img.imageWidth * 0.5)
            img_h = img_w / ratio
            if img_h > 120 * mm:
                img_h = 120 * mm
                img_w = img_h * ratio
            img.drawWidth = img_w
            img.drawHeight = img_h
            elements.append(img)
        except Exception:
            pass

    # Altbilgi (Artık create_report_canvas yapıyor ama boşluk veriyoruz)
    elements.append(Spacer(1, 10 * mm))

    doc.build(
        elements,
        onFirstPage=create_report_canvas,
        onLaterPages=create_report_canvas
    )
    pdf_bytes = buf.getvalue()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)

    return pdf_bytes


# ===================================================================
# Excel Rapor
# ===================================================================
def generate_excel_report(
    measurement_table: List[Dict],
    summary: Dict,
    calibration_info: Dict,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Excel (.xlsx) ölçüm raporu oluşturur.

    Returns:
        Excel dosyasının byte içeriği
    """
    wb = openpyxl.Workbook()

    # ── Ölçüm Sayfası ──
    ws = wb.active
    ws.title = "Ölçüm Sonuçları"

    # Stiller
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="3B82F6")
    subtitle_font = Font(name="Calibri", size=9, color="8B99AD")
    data_font = Font(name="Calibri", size=10)
    cap_font = Font(name="Calibri", size=10, color="3B82F6", bold=True)
    length_font = Font(name="Calibri", size=10, color="10B981", bold=True)
    border_thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Başlık
    ws.merge_cells("A1:D1")
    ws["A1"].value = "CNC Parça Ölçüm Raporu"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws.merge_cells("A2:D2")
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["A2"].value = f"Oluşturulma: {now}"
    ws["A2"].font = subtitle_font

    # Kalibrasyon bilgisi
    row = 4
    ws.cell(row=row, column=1, value="Kalibrasyon").font = Font(bold=True, size=11, color="3B82F6")
    row += 1
    ws.cell(row=row, column=1, value="Piksel/mm").font = data_font
    ws.cell(row=row, column=2, value=round(calibration_info.get("pixels_per_mm", 1.0), 4)).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="Profil Adı").font = data_font
    ws.cell(row=row, column=2, value=calibration_info.get("name", "Varsayılan")).font = data_font

    # Özet
    row += 2
    ws.cell(row=row, column=1, value="Ölçüm Özeti").font = Font(bold=True, size=11, color="3B82F6")
    row += 1
    if summary.get("total_sections", 0) > 0:
        for label, val in [
            ("Toplam Bölüm", str(summary["total_sections"])),
            ("Min Çap (mm)", f"{summary['min_diameter_mm']:.4f}"),
            ("Max Çap (mm)", f"{summary['max_diameter_mm']:.4f}"),
            ("Toplam Uzunluk (mm)", f"{summary['total_length_mm']:.4f}"),
        ]:
            ws.cell(row=row, column=1, value=label).font = data_font
            ws.cell(row=row, column=2, value=val).font = data_font
            row += 1

    # Ölçüm tablosu
    row += 1
    ws.cell(row=row, column=1, value="Ölçüm Tablosu").font = Font(bold=True, size=11, color="3B82F6")
    row += 1

    # Tablo başlıkları
    headers = ["ID", "Tip", "Açıklama", "Ölçülen (mm)", "Hedef", "Tol. (±)", "Durum"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_thin

    # Durum hücreleri için PatternFill boyama tipleri
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    pass_font = Font(name="Calibri", size=10, color="166534", bold=True)
    fail_font = Font(name="Calibri", size=10, color="991B1B", bold=True)

    # Tablo verileri
    for entry in measurement_table:
        row += 1
        ws.cell(row=row, column=1, value=entry["id"]).font = data_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = border_thin

        type_cell = ws.cell(row=row, column=2, value=entry["type"])
        type_cell.font = cap_font if entry["type"] == "Çap" else length_font
        type_cell.alignment = center_align
        type_cell.border = border_thin

        ws.cell(row=row, column=3, value=entry["description"]).font = data_font
        ws.cell(row=row, column=3).alignment = left_align
        ws.cell(row=row, column=3).border = border_thin

        ws.cell(row=row, column=4, value=round(entry["measured_mm"], 4)).font = data_font
        ws.cell(row=row, column=4).alignment = right_align
        ws.cell(row=row, column=4).number_format = "0.0000"
        ws.cell(row=row, column=4).border = border_thin
        
        # Hedef
        t_val = entry.get("target")
        ws.cell(row=row, column=5, value=t_val if t_val is not None else "-").font = data_font
        ws.cell(row=row, column=5).alignment = center_align
        ws.cell(row=row, column=5).border = border_thin
        
        # Tolerans
        tol_val = entry.get("tol")
        ws.cell(row=row, column=6, value=tol_val if tol_val is not None else "-").font = data_font
        ws.cell(row=row, column=6).alignment = center_align
        ws.cell(row=row, column=6).border = border_thin
        
        # Durum (PASS/FAIL)
        status = entry.get("status") or "-"
        status_cell = ws.cell(row=row, column=7, value=status)
        status_cell.alignment = center_align
        status_cell.border = border_thin
        
        if status == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif status == "FAIL":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.font = data_font

    # Kolon genişlikleri
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Kaydet
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(excel_bytes)

    return excel_bytes
